from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_non_ascii(file_path, sheet_name=None, save_path=None):
    # Load workbook
    wb = load_workbook(file_path)
    
    # Use specified sheet or active sheet
    ws = wb[sheet_name] if sheet_name else wb.active

    # Define highlight fill (light yellow)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate over all rows and cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Check if any character is outside ASCII range
                if any(ord(ch) > 127 for ch in cell.value):
                    cell.fill = highlight_fill

    # Save updated workbook
    save_path = save_path or file_path.replace(".xlsx", "_highlighted.xlsx")
    wb.save(save_path)
    print(f"File saved to {save_path}")

# Example usage:
highlight_non_ascii("data.xlsx")  # highlights in active sheet and saves as data_highlighted.xlsx